"""
============================================================
  Moodle VPL Student Attempts Extractor - FULL VERSION
  CSE-A Section | 17 Activities | Auto Lab Name Fetch
============================================================

SETUP (run once in terminal):
  pip install requests beautifulsoup4 openpyxl pandas

HOW TO RUN:
  python moodle_vpl_extractor_full.py

MODES:
  MODE 1 - Extract only CSE-A students (default, faster)
  MODE 2 - Extract ALL students across all activities
  Change EXTRACTION_MODE below to switch.
============================================================
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re, time, sys, os, tempfile
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ============================================================
#  >>>  CONFIG - ONLY SECTION YOU NEED TO EDIT  <<<
# ============================================================

MOODLE_BASE_URL       = "https://rulms.reva.edu.in"
MOODLE_SESSION_COOKIE = "41e64768a46aa80e1984a286075d753f"
OUTPUT_FILE           = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "CSE_A_VPL_Report.xlsx"
)
DELAY                 = 0.1     # seconds between requests

# MODE 1 = only CSE-A students below
# MODE 2 = all students found in each activity's submissions list
EXTRACTION_MODE = 1

# Prints the raw popup HTML headers/cells + fallback path for the first
# few attempts of the first few students of EVERY activity, so you can
# see exactly why "Grade" is/isn't being found. Turn off once fixed.
DEBUG_GRADE_EXTRACTION = False
DEBUG_GRADE_SAMPLE_SIZE = 0   # total debug prints across the whole run
_debug_attempts_logged = [0]  # mutable counter, don't edit


def save_workbook_safely(wb, path):
    """Save workbook to a temporary file and atomically replace the target.

    If the target file is locked on Windows, save to an alternate file in the
    same directory and return that alternate path.
    """
    target_dir = os.path.dirname(path) or os.getcwd()
    os.makedirs(target_dir, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(prefix="tmp_workbook_", suffix=".xlsx", dir=target_dir)
    os.close(fd)
    try:
        wb.save(tmp_path)
        try:
            os.replace(tmp_path, path)
            return path
        except PermissionError:
            alt_path = f"{path}.locked.xlsx"
            if os.path.exists(alt_path):
                os.remove(alt_path)
            os.replace(tmp_path, alt_path)
            return alt_path
    except Exception:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass

# VPL activity IDs from the provided Moodle submission-list URLs
VPL_ACTIVITIES = [
    {"id": "24261", "name": ""},
    {"id": "24262", "name": ""},
    {"id": "24263", "name": ""},
    {"id": "26228", "name": ""},
    {"id": "24264", "name": ""},
    {"id": "26229", "name": ""},
    {"id": "24265", "name": ""},
    {"id": "24266", "name": ""},
    {"id": "24267", "name": ""},
    {"id": "29812", "name": ""},
    {"id": "24268", "name": ""},
    {"id": "29813", "name": ""},
    {"id": "24269", "name": ""},
    {"id": "24270", "name": ""},
    {"id": "32909", "name": ""},
    {"id": "34177", "name": ""},
    {"id": "35145", "name": ""},
]


# CSE-A Student List (Roll No -> Name)
CSE_A_STUDENTS = {
    "R24EF308": "TEJAS G",
    "R25EF001": "AAGNEYA A PRESAD",
    "R25EF002": "ABDULLAH SUBHAAN K",
    "R25EF003": "ABHIK KUMAR BHUNIA",
    "R25EF004": "ABHINAV PRASAD E P",
    "R25EF005": "ABHISHEK B R",
    "R25EF006": "ABHISHEK CHALWA",
    "R25EF007": "ADARSH PATIL",
    "R25EF008": "ADARSH SHRIPAD PATIL",
    "R25EF009": "ADARSH SUDHEER GADAGIN",
    "R25EF010": "ADHIRAJ PARUTABAD",
    "R25EF011": "ADITI DEV",
    "R25EF012": "ADITTHYA S S",
    "R25EF013": "AGNELA PEARL VAS",
    "R25EF014": "AISHANI SEJAL",
    "R25EF015": "AISHVARRYA",
    "R25EF016": "AISHWARYA A J",
    "R25EF017": "AISHWARYA H K",
    "R25EF018": "AKASH S",
    "R25EF019": "AKHIL SATHISH KUMAR",
    "R25EF020": "AKRITI GUPTA",
    "R25EF021": "AKSHAY N",
    "R25EF022": "AKSHITHA N",
    "R25EF023": "ALFIYA NAAZ JAMADAR",
    "R25EF024": "AMAL SOORYA C",
    "R25EF025": "AMRITA JYOTI",
    "R25EF026": "AMRUTA SULIBHAVI",
    "R25EF027": "ANANYA SINGH",
    "R25EF028": "ANJANKUMAR B H",
    "R25EF029": "ANKAN KUMAR MANNA",
    "R25EF030": "ANKIT VIJAY",
    "R25EF031": "ANKITH KUMAR JHA",
    "R25EF032": "ANVITH E P",
    "R25EF033": "ARGHYAPRIYA SARKAR",
    "R25EF034": "ARUN KUMAR M",
    "R25EF035": "ARYAN PATILKULKARNI",
    "R25EF036": "ASHER ANIL GOVADA",
    "R25EF037": "ASHISH J",
    "R25EF038": "ASHUTOSH NARAYAN KULKARNI",
    "R25EF039": "ASHWIN K",
    "R25EF040": "ATHARVA R PADUKONE",
    "R25EF041": "AVANEESH K RAO",
    "R25EF042": "AVI SRIVASTAVA",
    "R25EF043": "AYUSH A KULKARNI",
    "R25EF044": "AYUSH RAJ",
    "R25EF045": "B M SHUBHANK",
    "R25EF046": "BASAVARAJ",
    "R25EF047": "BASAVARAJ KEMPANNA HEBBAL",
    "R25EF048": "BEKKAM SHANMUKA SAI",
    "R25EF049": "BHAVANA H R",
    "R25EF050": "BHAVASUDHAN S",
    "R25EF051": "BHOOMIKA J E",
    "R25EF052": "BHUVANESH D",
    "R25EF053": "BURRA ROHAN SABURI",
}

# ============================================================
#  STYLES
# ============================================================
DARK_BLUE  = "1F3864"
MED_BLUE   = "2B5592"
LIGHT_BLUE = "DAE8FC"
GREEN      = "E2EFDA"
DARK_GREEN = "1E6B1E"
YELLOW     = "FFF2CC"
WHITE      = "FFFFFF"
ORANGE     = "FCE4D6"

def _thin():
    s = Side(style="thin", color="B0B8C8")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, bg=DARK_BLUE, size=11):
    cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin()

def _dat(cell, row_idx, wrap=False, center=False, bg=None):
    fill = bg or (LIGHT_BLUE if row_idx % 2 == 0 else WHITE)
    cell.font      = Font(name="Calibri", size=10)
    cell.fill      = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap)
    cell.border    = _thin()


# ============================================================
#  SESSION
# ============================================================
def build_session():
    s = requests.Session()
    s.headers.update({"User-Agent":
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/124.0 Safari/537.36"})
    s.cookies.set("MoodleSession", MOODLE_SESSION_COOKIE,
                  domain="rulms.reva.edu.in", path="/")
    return s


def fetch(session, url, max_retries=5):
    """Fetch URL with automatic retry on timeout, connection issues, or HTML parsing failures."""
    for attempt in range(1, max_retries + 1):
        try:
            r = session.get(url, timeout=45)
            r.raise_for_status()
            html = r.text or ""
            return BeautifulSoup(html, "html.parser")
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError,
                requests.exceptions.RequestException) as e:
            wait = attempt * 5   # 5s, 10s, 15s, 20s, 25s
            print(f"\n  Network issue (attempt {attempt}/{max_retries}) - "
                  f"waiting {wait}s before retry...")
            time.sleep(wait)
            if attempt == max_retries:
                print(f"  SKIPPING after {max_retries} failed attempts: {url}")
                return BeautifulSoup("", "html.parser")
        except Exception as e:
            print(f"  Unexpected error fetching {url}: {e}")
            return BeautifulSoup("", "html.parser")
    return BeautifulSoup("", "html.parser")


def check_login(soup, context=""):
    page_text = soup.get_text(" ", strip=True).lower()
    title_text = ""
    title = soup.find("title")
    if title:
        title_text = title.get_text(" ", strip=True).lower()

    # A normal Moodle page can legitimately include a hidden login token field.
    # Only treat the page as a login-expired response when it truly presents a
    # login screen or the explicit logout message.
    if ("you are not logged in" in page_text or
            "log in to the site" in title_text or
            "login to the site" in title_text):
        print(f"\n  ERROR: Session cookie expired! {context}")
        print("  Get a fresh MoodleSession cookie from your browser")
        print("  and update MOODLE_SESSION_COOKIE in the CONFIG section.")
        sys.exit(1)


# ============================================================
#  AUTO-FETCH LAB NAMES
# ============================================================
def fetch_lab_names(session):
    """Fetch the actual lab/activity name from each VPL activity page."""
    print("\n  Fetching lab names from Moodle...")
    for activity in VPL_ACTIVITIES:
        if activity["name"]:
            continue
        url  = f"{MOODLE_BASE_URL}/mod/vpl/view.php?id={activity['id']}"
        soup = fetch(session, url)
        # Try <title> tag first
        title = soup.find("title")
        name  = ""
        if title:
            # Moodle titles are like "Lab Name: VPL | Course | Site"
            parts = title.get_text().split(":")
            name  = parts[0].strip() if parts else ""
        # Fallback: h2 heading
        if not name:
            h2 = soup.find("h2")
            if h2:
                name = h2.get_text(strip=True)
        # Fallback: page-header
        if not name:
            ph = soup.find(class_=re.compile("page-header|activity-name|instancename"))
            if ph:
                name = ph.get_text(strip=True)
        activity["name"] = name or f"Activity {activity['id']}"
        print(f"    id={activity['id']}  ->  {activity['name']}")
        time.sleep(0.3)
    print("  Lab names ready.")


# ============================================================
#  GET STUDENTS
# ============================================================
def parse_students_from_soup(soup):
    """
    Parse student rows from a submissions list page.
    Extracts: userid, full name from profile link, and other columns.
    """
    students = []
    for table in soup.find_all("table"):
        hrow = table.find("tr")
        if not hrow:
            continue
        headers = [c.get_text(strip=True).lower()
                   for c in hrow.find_all(["th", "td"])]
        # Must have at least one student-like header
        if not any(h in " ".join(headers)
                   for h in ["name", "student", "participant", "full"]):
            continue

        for row in table.find_all("tr")[1:]:
            cells = row.find_all(["td", "th"])
            if not cells:
                continue

            # Extract userid AND real name from profile link
            userid    = None
            real_name = ""
            for link in row.find_all("a", href=True):
                href = link["href"]
                # Profile link contains the real name as link text
                m = re.search(r"/user/view\.php\?id=(\d+)", href)
                if m:
                    userid    = m.group(1)
                    real_name = link.get_text(strip=True)
                    break
                # VPL submission link
                m2 = re.search(r"userid=(\d+)", href)
                if m2 and not userid:
                    userid = m2.group(1)

            # If no profile link found, try userid from any link
            if not userid:
                for link in row.find_all("a", href=True):
                    m = re.search(r"userid=(\d+)", link["href"])
                    if m:
                        userid = m.group(1)
                        break

            # Get name: prefer profile link text, fallback to first cell
            if not real_name:
                # Check first cell for a profile link
                first_cell = cells[0] if cells else None
                if first_cell:
                    pl = first_cell.find("a", href=True)
                    if pl:
                        real_name = pl.get_text(strip=True)
                    else:
                        real_name = first_cell.get_text(strip=True)

            # Skip rows where name is just a number (row index artifact)
            if real_name and real_name.isdigit():
                real_name = ""

            entry = {"name": real_name, "userid": userid or ""}
            for i, h in enumerate(headers[1:], 1):
                if i < len(cells):
                    val = cells[i].get_text(strip=True)
                    entry[h] = val

            if userid:
                students.append(entry)

    return students


def get_all_students_from_activity(session, activity_id):
    """
    Fetch ALL students using the exact URL pattern from
    the "Show all XXXX" button:
    submissionslist.php?id=X&showgrades=0&group=-1&tilast&tifirst&tperpage=5000&thiddenfields
    """
    # Use the exact URL pattern the Show All button generates
    url = (f"{MOODLE_BASE_URL}/mod/vpl/views/submissionslist.php"
           f"?id={activity_id}&showgrades=0&group=-1"
           f"&tilast&tifirst&tperpage=5000&thiddenfields")

    soup = fetch(session, url)
    check_login(soup, f"activity id={activity_id}")
    students  = parse_students_from_soup(soup)
    seen_uids = {s["userid"] for s in students}

    # Fallback: if still paginated, loop through pages
    if len(students) < 50:
        page = 0
        while True:
            page_url = (f"{MOODLE_BASE_URL}/mod/vpl/views/submissionslist.php"
                        f"?id={activity_id}&showgrades=0&group=-1"
                        f"&tilast&tifirst&tperpage=100&thiddenfields"
                        f"&tpage={page}")
            psoup = fetch(session, page_url)
            found = parse_students_from_soup(psoup)
            if not found:
                break
            new_count = 0
            for s in found:
                if s["userid"] not in seen_uids:
                    students.append(s)
                    seen_uids.add(s["userid"])
                    new_count += 1
            if new_count == 0:
                break
            page += 1
            time.sleep(0.3)

    return students


def filter_cse_a(students):
    """
    Match fetched Moodle students against CSE-A list.
    Uses 4 levels of matching - most strict to most fuzzy:
      1. Exact match (both names identical)
      2. Full containment (one name contains the other)
      3. Token overlap (2+ words in common)
      4. First-name + last-initial match
    """
    matched   = []
    unmatched = list(CSE_A_STUDENTS.items())
    used_uids = set()

    def normalize(s):
        return re.sub(r'\s+', ' ', s.strip().upper())

    def tokens(s):
        return set(normalize(s).split())

    def first_name(s):
        parts = normalize(s).split()
        return parts[0] if parts else ""

    def last_token(s):
        parts = normalize(s).split()
        return parts[-1] if parts else ""

    # Build index for speed
    cse_list = [(roll, name, normalize(name), tokens(name))
                for roll, name in CSE_A_STUDENTS.items()]

    for s in students:
        if s["userid"] in used_uids:
            continue
        moodle_norm   = normalize(s["name"])
        moodle_tokens = tokens(s["name"])
        best_match    = None
        best_score    = 0

        for roll, cse_name, cse_norm, cse_tokens in cse_list:
            # Already matched this roll number
            if not any(r == roll for r, n in unmatched):
                continue

            score = 0

            # Level 1: Exact
            if moodle_norm == cse_norm:
                score = 100

            # Level 2: Full containment
            elif cse_norm in moodle_norm or moodle_norm in cse_norm:
                score = 90

            # Level 3: Token overlap - count shared words
            else:
                common = moodle_tokens & cse_tokens
                # Remove very short tokens (initials like A, B, S)
                common = {t for t in common if len(t) > 1}
                if len(common) >= 2:
                    score = 70 + len(common)
                elif len(common) == 1 and len(list(common)[0]) >= 4:
                    # One long word in common (e.g. ASHUTOSH)
                    score = 60

            # Level 4: First name matches + last token matches
            if score == 0:
                if (first_name(s["name"]) == first_name(cse_name) and
                        last_token(s["name"]) == last_token(cse_name)):
                    score = 55

            # Level 5: First name matches alone (weak, only if unique)
            if score == 0:
                if first_name(s["name"]) == first_name(cse_name):
                    score = 30

            if score > best_score:
                best_score = score
                best_match = (roll, cse_name)

        # Accept match if score is good enough
        if best_match and best_score >= 55:
            roll, cse_name = best_match
            matched.append({
                "name":    s["name"],
                "userid":  s["userid"],
                "roll_no": roll,
                "section": "CSE-A",
            })
            unmatched  = [(r, n) for r, n in unmatched if r != roll]
            used_uids.add(s["userid"])

    return matched, unmatched


# ============================================================
#  GET ATTEMPTS
# ============================================================
def extract_lab_marks(grade_str):
    """Parse various grade formats to return the student's numeric mark.

    Examples handled:
      - '10 / 10'  -> 10
      - '8/10'     -> 8
      - '8 out of 10' or '8 of 10' -> 8
      - '8'        -> 8
      - ''         -> ''
    """
    if not grade_str:
        return ""
    s = str(grade_str).strip()
    # X / Y or X/Y
    m = re.search(r'(\d+(?:\.\d+)?)\s*/\s*\d+(?:\.\d+)?', s)
    if m:
        val = float(m.group(1))
        return int(val) if val == int(val) else val
    # 'X out of Y' or 'X of Y'
    m2 = re.search(r'(\d+(?:\.\d+)?)\s*(?:out of|of)\s*\d+(?:\.\d+)?', s, re.I)
    if m2:
        val = float(m2.group(1))
        return int(val) if val == int(val) else val
    # plain number '8' or '8.5'
    m3 = re.match(r'^(\d+(?:\.\d+)?)$', s)
    if m3:
        val = float(m3.group(1))
        return int(val) if val == int(val) else val
    return ""


def fetch_student_name(session, userid):
    """Fetch real student name from their Moodle profile page."""
    try:
        url  = f"{MOODLE_BASE_URL}/user/view.php?id={userid}"
        soup = fetch(session, url)
        # Try h1 page-header first
        for tag in ["h1", "h2"]:
            el = soup.find(tag)
            if el:
                name = el.get_text(strip=True)
                if name and not name.isdigit() and len(name) > 2:
                    return name
        # Try page title
        title = soup.find("title")
        if title:
            parts = title.get_text().split(":")
            name  = parts[0].strip()
            if name and not name.isdigit():
                return name
    except Exception:
        pass
    return ""


def fetch_grade_from_submission_view(session, url):
    """Fetch grade text from the submission view page if not present in the attempts table."""
    if not url:
        return ""
    try:
        soup = fetch(session, url)

        # 1) Look for a dedicated grade input/select field first -- VPL's
        #    submission view often renders the grade as a form field
        #    (<input name="grade" value="10.00">) rather than plain text,
        #    which get_text() would miss entirely.
        for inp in soup.find_all("input"):
            name = (inp.get("name") or "").lower()
            if "grade" in name and inp.get("value"):
                v = inp["value"].strip()
                if re.match(r'^[0-9]+(\.[0-9]+)?$', v):
                    return v
        for sel in soup.find_all("select"):
            name = (sel.get("name") or "").lower()
            if "grade" in name:
                chosen = sel.find("option", selected=True)
                if chosen and chosen.get("value"):
                    v = chosen["value"].strip()
                    if re.match(r'^[0-9]+(\.[0-9]+)?$', v):
                        return v

        # 2) Look for elements Moodle/VPL commonly tags with a grade class.
        for el in soup.find_all(class_=re.compile(r'grade|vpl.?grade', re.I)):
            t = el.get_text(" ", strip=True)
            m = re.search(r'([0-9]+(?:\.[0-9]+)?\s*/\s*[0-9]+(?:\.[0-9]+)?)', t)
            if m:
                return m.group(1)

        # 3) Text-pattern search across the whole page.
        text = soup.get_text(" ", strip=True)
        patterns = [
            r'Grade\s*[:]\s*([0-9]+(?:\.[0-9]+)?\s*/\s*[0-9]+(?:\.[0-9]+)?)',
            r'Proposed grade\s*[:]\s*([0-9]+(?:\.[0-9]+)?\s*/\s*[0-9]+(?:\.[0-9]+)?)',
            r'Grade\s*[:]\s*([0-9]+(?:\.[0-9]+)?)',
            r'Proposed grade\s*[:]\s*([0-9]+(?:\.[0-9]+)?)',
        ]
        for pat in patterns:
            m = re.search(pat, text, re.I)
            if m:
                return m.group(1)
        # fallback to any simple X / Y pattern
        m = re.search(r'([0-9]+(?:\.[0-9]+)?\s*/\s*[0-9]+(?:\.[0-9]+)?)', text)
        if m:
            return m.group(1)

        if DEBUG_GRADE_EXTRACTION and _debug_attempts_logged[0] < DEBUG_GRADE_SAMPLE_SIZE:
            snippet = text[:300].replace("\n", " ")
            print(f"\n    [DEBUG submission_view] NO grade pattern found at {url}")
            print(f"    [DEBUG submission_view] page text starts: {snippet!r}")
    except Exception as e:
        if DEBUG_GRADE_EXTRACTION:
            print(f"\n    [DEBUG submission_view] error fetching {url}: {e}")
    return ""


def get_attempts(session, activity_id, userid, student_name):
    """Scrape all attempts for one student in one activity."""
    url = (f"{MOODLE_BASE_URL}/mod/vpl/views/previoussubmissionslist.php"
           f"?id={activity_id}&userid={userid}&inpopup=1")
    soup = fetch(session, url)

    def extract_grade_from_cell(cell):
        txt = cell.get_text(strip=True)
        # Direct grade pattern X / Y
        if re.search(r'\d+\s*/\s*\d+', txt):
            return txt
        # Check title attribute
        for tag in cell.find_all(attrs={"title": True}):
            t = tag["title"].strip()
            if re.search(r'\d+\s*/\s*\d+', t):
                return t
        # Check any nested span/div for grade
        for span in cell.find_all(["span", "div"]):
            st = span.get_text(strip=True)
            if re.search(r'\d+\s*/\s*\d+', st):
                return st
        return txt if txt not in ("", "-", "--", "--") else ""

    def pick(raw, *candidates):
        rl = {k.lower(): v for k, v in raw.items()}
        for k in candidates:
            if k.lower() in rl:
                return rl[k.lower()]
        for k in candidates:
            for rk, rv in rl.items():
                if k.lower() in rk and "name" not in rk:
                    return rv
        return ""

    attempts = []
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue
        raw_headers = [c.get_text(strip=True)
                       for c in rows[0].find_all(["th", "td"])]
        if not raw_headers:
            continue

        for n, row in enumerate(rows[1:], 1):
            cells = row.find_all(["td", "th"])
            if not cells:
                continue
            raw       = {}
            raw_cells = {}
            for j, cell in enumerate(cells):
                key = raw_headers[j] if j < len(raw_headers) else f"col_{j}"
                raw[key]       = cell.get_text(strip=True)
                raw_cells[key] = cell

            sub_url = ""
            sub_id  = ""
            for link in row.find_all("a", href=True):
                href = link["href"]
                if "submissionview" in href or "submissionid" in href:
                    sub_url = (href if href.startswith("http")
                               else MOODLE_BASE_URL + href)
                    m = re.search(r"submissionid=(\d+)", href)
                    if m:
                        sub_id = m.group(1)
                    break

            grade = pick(raw, "grade", "mark", "score", "result")
            if not grade or not re.search(r'\d', grade):
                for key, cell in raw_cells.items():
                    g = extract_grade_from_cell(cell)
                    if g and re.search(r'\d+\s*/\s*\d+', g):
                        grade = g
                        break

            grade_source = "popup" if (grade and re.search(r'\d', grade)) else ""

            if (not grade or not re.search(r'\d', grade)) and sub_url:
                # Popup table had no usable grade -- always try the
                # submission's own page, since VPL usually renders the
                # grade there even when the popup list doesn't.
                fallback_grade = fetch_grade_from_submission_view(session, sub_url)
                if fallback_grade:
                    grade = fallback_grade
                    grade_source = "submission_view"

            if DEBUG_GRADE_EXTRACTION and _debug_attempts_logged[0] < DEBUG_GRADE_SAMPLE_SIZE:
                _debug_attempts_logged[0] += 1
                print(f"\n    [DEBUG attempt#{n}] popup headers={raw_headers}")
                print(f"    [DEBUG attempt#{n}] popup row cells={raw}")
                print(f"    [DEBUG attempt#{n}] sub_url={sub_url or '(none found)'}")
                print(f"    [DEBUG attempt#{n}] final grade={grade!r}  source={grade_source or 'NONE FOUND'}")

            attempts.append({
                "Student Name":           student_name,
                "User ID":                userid,
                "Class Section":          "",   # filled later manually
                "Attempt #":              str(n),
                "Submission Date & Time": pick(raw, "date", "submission date", "time"),
                "Description":            pick(raw, "description", "desc", "file"),
                "Grade":                  grade,
                "Lab Marks":              extract_lab_marks(grade),
                "Status":                 pick(raw, "status", "state"),
                "Submission View URL":    sub_url,
                "Submission ID":          sub_id,
            })

    if not attempts:
        txt = soup.get_text(separator=" ").lower()
        status = ("No submissions"
                  if "no submission" in txt or "no attempt" in txt
                  else "Not accessible" if "log in" in txt
                  else "No data")
        attempts = [{
            "Student Name": student_name, "User ID": userid,
            "Class Section": "", "Attempt #": "0",
            "Submission Date & Time": "", "Description": "",
            "Grade": "", "Lab Marks": "", "Status": status,
            "Submission View URL": "", "Submission ID": "",
        }]
    return attempts


# ============================================================
#  EXCEL SHEETS
# ============================================================
ALL_ATT_COLS   = [
    "Student Name", "User ID", "Class Section", "Attempt #",
    "Submission Date & Time", "Description", "Grade", "Lab Marks",
    "Status", "Submission View URL", "Submission ID"
]
ALL_ATT_WIDTHS = [24, 10, 12, 10, 22, 26, 12, 10, 14, 55, 14]

SUM_COLS   = [
    "Roll No", "Student Name", "User ID", "Class Section",
    "Attempt Count", "Attempt Numbers", "Attempt Marks",
    "Submission Dates & Times",
    "Marks", "Lab Marks", "Latest Submission", "Description"
]
SUM_WIDTHS = [12, 26, 10, 12, 14, 36, 26, 48, 28, 10, 22, 48]


def write_attempts_sheet(wb, sheet_name, all_attempts):
    ws = wb.create_sheet(sheet_name[:31])
    for ci, h in enumerate(ALL_ATT_COLS, 1):
        _hdr(ws.cell(1, ci, h))
    ws.row_dimensions[1].height = 28
    for ri, att in enumerate(all_attempts, 2):
        for ci, col in enumerate(ALL_ATT_COLS, 1):
            val = att.get(col, "") or ""
            c   = ws.cell(ri, ci, val)
            is_url    = col == "Submission View URL"
            is_center = col in ("User ID", "Attempt #", "Grade",
                                "Lab Marks", "Status", "Submission ID",
                                "Class Section")
            _dat(c, ri, center=is_center)
            if is_url and val:
                c.hyperlink = val
                c.font = Font(name="Calibri", size=10,
                              color="0563C1", underline="single")
    for ci, w in enumerate(ALL_ATT_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def write_summary_sheet(wb, sheet_name, all_attempts, students):
    ws = wb.create_sheet(sheet_name[:31])
    for ci, h in enumerate(SUM_COLS, 1):
        _hdr(ws.cell(1, ci, h), bg=MED_BLUE)
    ws.row_dimensions[1].height = 30

    # uid -> ordered list of attempt records (num, date, grade, lab_mark, desc)
    agg = defaultdict(list)
    for att in all_attempts:
        uid  = str(att.get("User ID", ""))
        anum = att.get("Attempt #", "0")
        if anum == "0":
            continue
        agg[uid].append({
            "num":      anum,
            "date":     att.get("Submission Date & Time", ""),
            "grade":    att.get("Grade", "").strip(),
            "lab_mark": att.get("Lab Marks", ""),
            "desc":     att.get("Description", ""),
        })

    for ri, s in enumerate(students, 2):
        uid     = str(s.get("userid", ""))
        records = agg[uid]

        attempt_nums  = ", ".join(r["num"] for r in records)
        # Attempt Marks lines up 1:1 with Attempt Numbers ("-" where no mark)
        attempt_marks = ", ".join(
            (str(r["lab_mark"]) if r["lab_mark"] != "" else "-")
            for r in records
        )
        dates_str     = ", ".join(r["date"] for r in records if r["date"])
        marks_str     = ", ".join(
            r["grade"] for r in records if r["grade"] and re.search(r'\d', r["grade"])
        )
        desc_str      = ", ".join(r["desc"] for r in records if r["desc"])
        latest        = records[-1]["date"] if records else ""
        attempt_count = len(records)
        lab_marks_val = (records[-1]["lab_mark"]
                          if records and records[-1]["lab_mark"] != "" else "")
        roll_no       = s.get("roll_no", "")

        row_vals = [
            roll_no, s.get("name", ""), uid, s.get("section", ""),
            attempt_count, attempt_nums, attempt_marks, dates_str,
            marks_str, lab_marks_val, latest, desc_str,
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(ri, ci, val)
            wrap_cols   = {6, 7, 8, 9, 12}
            center_cols = {3, 4, 5, 10}
            _dat(c, ri - 1,
                 wrap=(ci in wrap_cols),
                 center=(ci in center_cols))

        max_items = max(
            dates_str.count(",") + 1 if dates_str else 1,
            desc_str.count(",")  + 1 if desc_str  else 1,
        )
        ws.row_dimensions[ri].height = min(15 * max(1, max_items // 3 + 1), 120)

    for ci, w in enumerate(SUM_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def write_attempt_marks_grid(wb, all_lab_data, students):
    """
    One row per Student + Activity, with a separate column for each
    attempt's mark (Attempt 1, Attempt 2, ... Attempt N), plus Best
    Mark and Total Attempts. N is the max attempt count seen across
    all students/activities.
    """
    ws = wb.create_sheet("Attempt Marks Grid"[:31])
    lab_names = list(all_lab_data.keys())

    # (uid, lab_name) -> [(attempt_num_int, lab_mark), ...] sorted by attempt #
    marks_map = defaultdict(list)
    for lab_name, attempts in all_lab_data.items():
        for att in attempts:
            uid  = str(att.get("User ID", ""))
            anum = att.get("Attempt #", "0")
            if anum == "0":
                continue
            try:
                anum_int = int(anum)
            except (ValueError, TypeError):
                anum_int = 0
            marks_map[(uid, lab_name)].append((anum_int, att.get("Lab Marks", "")))
    for k in marks_map:
        marks_map[k].sort(key=lambda x: x[0])

    max_attempts = max((len(v) for v in marks_map.values()), default=1)
    max_attempts = max(max_attempts, 1)

    cols = (["Roll No", "Student Name", "User ID", "Class Section", "Activity"]
            + [f"Attempt {i}" for i in range(1, max_attempts + 1)]
            + ["Best Mark", "Total Attempts"])
    for ci, h in enumerate(cols, 1):
        _hdr(ws.cell(1, ci, h), bg=DARK_GREEN)
    ws.row_dimensions[1].height = 28

    ri = 2
    for s in students:
        uid     = str(s.get("userid", ""))
        roll_no = s.get("roll_no", "")
        name    = s.get("name", "")
        section = s.get("section", "")

        for lab_name in lab_names:
            records = marks_map.get((uid, lab_name), [])
            row_vals = [roll_no, name, uid, section, lab_name]

            numeric_marks = []
            for i in range(max_attempts):
                if i < len(records):
                    mark = records[i][1]
                    row_vals.append(mark if mark != "" else "-")
                    if mark != "":
                        try:
                            numeric_marks.append(float(mark))
                        except (ValueError, TypeError):
                            pass
                else:
                    row_vals.append("")

            if numeric_marks:
                best_val = max(numeric_marks)
                best = int(best_val) if best_val == int(best_val) else best_val
            else:
                best = ""
            row_vals.append(best)
            row_vals.append(len(records))

            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(ri, ci, val)
                is_attempt_col = 6 <= ci <= 5 + max_attempts
                bg = GREEN if is_attempt_col and val not in ("", "-") else None
                bg = ORANGE if is_attempt_col and val == "" else bg
                _dat(c, ri - 1, center=(ci >= 3), bg=bg)
            ri += 1

    widths = [12, 26, 10, 12, 24] + [11] * max_attempts + [11, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "C2"


def write_per_activity_attempt_marks(wb, all_lab_data, students):
    """Create one worksheet per activity showing every student's attempt marks."""
    for lab_name, attempts in all_lab_data.items():
        ws = wb.create_sheet((lab_name[:24] + " - Attempt Marks")[:31])

        marks_map = defaultdict(list)
        for att in attempts:
            uid = str(att.get("User ID", ""))
            anum = att.get("Attempt #", "0")
            if anum == "0":
                continue
            try:
                anum_int = int(anum)
            except (ValueError, TypeError):
                anum_int = 0
            marks_map[uid].append((anum_int, att.get("Lab Marks", "")))

        for uid in marks_map:
            marks_map[uid].sort(key=lambda x: x[0])

        max_attempts = max((len(v) for v in marks_map.values()), default=1)
        max_attempts = max(max_attempts, 1)

        cols = ["Roll No", "Student Name", "User ID", "Class Section"]
        cols += [f"Attempt {i}" for i in range(1, max_attempts + 1)]
        cols += ["Best Mark", "Total Attempts"]

        for ci, h in enumerate(cols, 1):
            _hdr(ws.cell(1, ci, h), bg=DARK_GREEN)
        ws.row_dimensions[1].height = 28

        ri = 2
        for s in students:
            uid = str(s.get("userid", ""))
            records = marks_map.get(uid, [])
            roll_no = s.get("roll_no", "")
            name = s.get("name", "")
            section = s.get("section", "")

            row_vals = [roll_no, name, uid, section]
            numeric_marks = []
            for i in range(max_attempts):
                if i < len(records):
                    mark = records[i][1]
                    row_vals.append(mark if mark != "" else "-")
                    if mark != "":
                        try:
                            numeric_marks.append(float(mark))
                        except (ValueError, TypeError):
                            pass
                else:
                    row_vals.append("")

            if numeric_marks:
                best_val = max(numeric_marks)
                best = int(best_val) if best_val == int(best_val) else best_val
            else:
                best = ""
            row_vals.append(best)
            row_vals.append(len(records))

            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(ri, ci, val)
                is_attempt_col = 5 <= ci <= 4 + max_attempts
                bg = GREEN if is_attempt_col and val not in ("", "-") else None
                bg = ORANGE if is_attempt_col and val == "" else bg
                _dat(c, ri - 1, center=(ci >= 3), bg=bg)
            ri += 1

        widths = [12, 26, 10, 12] + [11] * max_attempts + [11, 14]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "E2"


def write_master_summary(wb, all_lab_data, students):
    """
    One row per student, one column per lab = Lab Marks grid.
    """
    ws = wb.create_sheet("Master Summary")
    lab_names   = [a["name"] for a in VPL_ACTIVITIES if a["name"]]
    master_cols = (["Roll No", "Student Name", "User ID", "Class Section"]
                   + lab_names
                   + ["Total Marks", "Labs Attempted"])

    for ci, h in enumerate(master_cols, 1):
        _hdr(ws.cell(1, ci, h), bg=DARK_GREEN)
    ws.row_dimensions[1].height = 30

    # Build marks lookup: uid -> {lab_name -> {"all": "5, 8, 10", "latest": 10}}
    marks_lookup = defaultdict(dict)
    for lab_name, attempts in all_lab_data.items():
        uid_marks = defaultdict(list)
        for att in attempts:
            uid = str(att.get("User ID", ""))
            lm  = att.get("Lab Marks", "")
            if lm != "" and att.get("Attempt #", "0") != "0":
                uid_marks[uid].append(lm)
        for uid, lms in uid_marks.items():
            marks_lookup[uid][lab_name] = {
                "all":    ", ".join(str(x) for x in lms),
                "latest": lms[-1] if lms else "",
            }

    for ri, s in enumerate(students, 2):
        uid     = str(s.get("userid", ""))
        roll_no = s.get("roll_no", "")
        total   = 0
        labs_done = 0

        row_vals = [roll_no, s.get("name", ""), uid, s.get("section", "")]
        for lab_name in lab_names:
            entry = marks_lookup[uid].get(lab_name, {"all": "", "latest": ""})
            # Show every attempt's mark for this lab (e.g. "5, 8, 10")
            row_vals.append(entry["all"])
            latest = entry["latest"]
            if latest != "":
                try:
                    total     += float(latest)
                    labs_done += 1
                except (ValueError, TypeError):
                    pass

        row_vals.append(int(total) if total == int(total) else total)
        row_vals.append(f"{labs_done}/{len(lab_names)}")

        for ci, val in enumerate(row_vals, 1):
            c  = ws.cell(ri, ci, val)
            is_lab_col = 5 <= ci <= 4 + len(lab_names)
            bg = GREEN if is_lab_col and val != "" else None
            bg = ORANGE if is_lab_col and val == "" else bg
            _dat(c, ri - 1, center=(ci >= 3), bg=bg, wrap=is_lab_col)
        ws.row_dimensions[ri].height = 18

    # Column widths
    for ci, w in enumerate([12, 28, 10, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ci in range(5, len(master_cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    ws.freeze_panes = "E2"


# ============================================================
#  UNMATCHED STUDENTS SHEET
# ============================================================
def write_unmatched_sheet(wb, unmatched_per_lab):
    ws = wb.create_sheet("Unmatched Students")
    for ci, h in enumerate(["Roll No", "CSE-A Name", "Activity", "Note"], 1):
        _hdr(ws.cell(1, ci, h), bg="C00000")
    ws.row_dimensions[1].height = 24

    ri = 2
    for lab_name, unmatched in unmatched_per_lab.items():
        for roll, name in unmatched:
            ws.cell(ri, 1, roll).border  = _thin()
            ws.cell(ri, 1).font          = Font(name="Calibri", size=10)
            ws.cell(ri, 1).fill          = PatternFill("solid", fgColor=ORANGE)
            ws.cell(ri, 2, name).border  = _thin()
            ws.cell(ri, 2).font          = Font(name="Calibri", size=10)
            ws.cell(ri, 2).fill          = PatternFill("solid", fgColor=ORANGE)
            ws.cell(ri, 3, lab_name).border = _thin()
            ws.cell(ri, 3).font             = Font(name="Calibri", size=10)
            ws.cell(ri, 3).fill             = PatternFill("solid", fgColor=ORANGE)
            ws.cell(ri, 4, "Not found in Moodle submissions list").border = _thin()
            ws.cell(ri, 4).font = Font(name="Calibri", size=10, italic=True, color="555555")
            ws.cell(ri, 4).fill = PatternFill("solid", fgColor=ORANGE)
            ri += 1

    for ci, w in enumerate([14, 30, 26, 38], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ============================================================
#  MAIN
# ============================================================
def fetch_attempts_for_student(activity_id, student, index, total):
    session = build_session()
    uid = student["userid"]
    name = student["name"]
    if not name or name.strip().isdigit():
        fetched = fetch_student_name(session, uid)
        if fetched:
            name = fetched
            student["name"] = fetched
    print(f"    [{index:>3}/{total}] {name}  (uid={uid})", end="  ", flush=True)

    ats = []
    for retry in range(3):
        ats = get_attempts(session, activity_id, uid, name)
        if ats:
            break
        print(f" retry {retry+1}..", end="", flush=True)
        time.sleep(3)

    real = len([a for a in ats if a.get("Attempt #", "0") != "0"])
    student["total_attempts"] = real
    print(f"-> {real} attempt(s)")
    return student, ats


def main():
    print("=" * 65)
    print("  Moodle VPL Extractor - CSE-A | All Labs")
    print(f"  Mode      : {'CSE-A students only' if EXTRACTION_MODE == 1 else 'All students'}")
    print(f"  Activities: {len(VPL_ACTIVITIES)}")
    print(f"  Mode      : Fetching ALL students with Moodle names as-is")
    print(f"  Started   : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65)

    session = build_session()

    # Step 1: Auto-fetch lab names
    fetch_lab_names(session)

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_lab_attempts  = {}   # lab_name -> list of attempts
    all_lab_students  = {}   # lab_name -> list of student dicts
    unmatched_per_lab = {}   # lab_name -> list of (roll, name)
    master_students   = []   # final unified student list

    # Step 2: Loop through each activity
    for lab_idx, activity in enumerate(VPL_ACTIVITIES, 1):
        act_id   = activity["id"]
        act_name = activity["name"]

        print(f"\n{'-'*65}")
        print(f"  [{lab_idx}/{len(VPL_ACTIVITIES)}] {act_name}  (id={act_id})")
        print(f"{'-'*65}")

        # Get ALL students from this activity (bypass pagination)
        print("  Fetching full student list (bypassing pagination)...")
        all_students = get_all_students_from_activity(session, act_id)
        print(f"  Total students in activity: {len(all_students)}")

        # Filter to CSE-A only if Mode 1
        if EXTRACTION_MODE == 1:
            students, unmatched = filter_cse_a(all_students)
            print(f"  CSE-A matched: {len(students)}  |  Unmatched: {len(unmatched)}")
            if unmatched:
                print(f"  Unmatched students: {[n for _, n in unmatched]}")
            unmatched_per_lab[act_name] = unmatched
        else:
            students  = all_students
            unmatched = []

        if not students:
            print(f"  WARNING: No CSE-A students matched for {act_name}.")
            if all_students:
                sample = [s["name"] for s in all_students[:5]]
                print(f"  Sample Moodle names in this activity: {sample}")
                print(f"  -> These may be a different batch/section.")
                print(f"  -> Skipping this activity for CSE-A.")
            continue

        # Build master student list from first activity
        if not master_students:
            master_students = students

        # Step 3: Get attempts for each student
        all_attempts = []
        print(f"  Fetching attempts for {len(students)} student(s)...")

        with ThreadPoolExecutor(max_workers=1) as executor:
            futures = {
                executor.submit(fetch_attempts_for_student, act_id, s, i, len(students)): i
                for i, s in enumerate(students, 1)
            }
            for future in as_completed(futures):
                _, ats = future.result()
                all_attempts.extend(ats)

        time.sleep(DELAY)

        real_total = len([a for a in all_attempts
                          if a.get("Attempt #", "0") != "0"])
        print(f"  Total attempts: {real_total}")

        # Step 4: Write sheets
        safe = re.sub(r'[\\/*?:\[\]]', '', act_name)[:24]
        write_attempts_sheet(wb, f"{safe} - Attempts", all_attempts)
        write_summary_sheet(wb,  f"{safe} - Summary",  all_attempts, students)
        print(f"  Sheets written: '{safe} - Attempts'  +  '{safe} - Summary'")

        os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
        wb_temp = openpyxl.Workbook()
        wb_temp.remove(wb_temp.active)
        write_attempts_sheet(wb_temp, f"{safe} - Attempts", all_attempts)
        write_summary_sheet(wb_temp, f"{safe} - Summary", all_attempts, students)
        wb_temp.save(f"CHECKPOINT_{act_id}.xlsx")
        print(f"  [Checkpoint] Saved CHECKPOINT_{act_id}.xlsx")

        saved_path = save_workbook_safely(wb, OUTPUT_FILE)
        print(f"  [Progress] Saved workbook: {saved_path}")
        if saved_path != OUTPUT_FILE:
            print(f"  [Warning] Original file locked; progress saved to {saved_path}")

        all_lab_attempts[act_name] = all_attempts
        all_lab_students[act_name] = students

    # Step 5: Master Summary
    if master_students and all_lab_attempts:
        print(f"\n{'-'*65}")
        print("  Building Master Summary (all students x all labs)...")
        write_master_summary(wb, all_lab_attempts, master_students)
        print("  Master Summary written.")

        print("  Building Attempt Marks Grid (per-attempt marks per activity)...")
        write_attempt_marks_grid(wb, all_lab_attempts, master_students)
        print("  Attempt Marks Grid written.")

        print("  Building per-activity attempt-mark sheets for all students...")
        write_per_activity_attempt_marks(wb, all_lab_attempts, master_students)
        print("  Per-activity attempt-mark sheets written.")

    # Step 6: Unmatched students sheet
    any_unmatched = any(v for v in unmatched_per_lab.values())
    if any_unmatched:
        write_unmatched_sheet(wb, unmatched_per_lab)
        print("  Unmatched Students sheet written.")

    # Save
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    final_path = save_workbook_safely(wb, OUTPUT_FILE)

    print(f"\n{'='*65}")
    if final_path != OUTPUT_FILE:
        print(f"  WARNING: Could not overwrite locked file. Saved workbook to: {final_path}")
        print(f"  Close the file '{OUTPUT_FILE}' in Excel and rename the alternate file if needed.")
    print(f"  DONE!")
    print(f"  Saved    : {OUTPUT_FILE}")
    print(f"  Sheets   : {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"             * {s}")
    print(f"  Finished : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()
